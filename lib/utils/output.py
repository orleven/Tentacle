#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @author: 'orleven'

import re
import sys
import logging
# from lib.core.data import kb
from lib.core.data import logger
from lib.core.log import LOGGER_HANDLER
from lib.core.settings import BANNER
from lib.core.settings import IS_WIN
from openpyxl import Workbook
from thirdparty.termcolor.termcolor import colored
from thirdparty.colorama.initialise import init as coloramainit

def banner():
    if IS_WIN:
        coloramainit()
    data_to_stdout(BANNER)

def output_excal(datalines,file,taskname = None):
    filename = file + '.xlsx'
    if taskname:
        logger.info('Task export to %s: %s' % (filename,taskname))
    else:
        logger.info('Export to %s...' % (filename))
    book = Workbook()
    ws = book.active
    i = 1
    titleList = []
    for line in datalines:
        i = i + 1
        for key in line:
            if key not in titleList:
                titleList.append(key)
                ws.cell(row=1, column=len(titleList)).value = key
            try:
                ws.cell(row=i, column=titleList.index(key) + 1).value = str(line[key])
                # if isinstance(line[key], int) or isinstance(line[key], str):
                #     ws.cell(row=i, column=titleList.index(key) + 1).value = line[key]
                # elif isinstance(line[key], list):
                #     ws.cell(row=i, column=titleList.index(key) + 1).value = str(line[key])
                # elif isinstance(line[key], dict):
                #     ws.cell(row=i, column=titleList.index(key) + 1).value = str(line[key])
                # elif isinstance(line[key], None):
                #     ws.cell(row=i, column=titleList.index(key) + 1).value = ""
                # else:
                #     ws.cell(row=i, column=titleList.index(key) + 1).value = str(line[key])
            except:
                ws.cell(row=i, column=titleList.index(key) + 1).value = "Some error."
    book.save(filename)
    if taskname:
        logger.sysinfo('Task exported to %s successful: %s' % (filename,taskname))
    else:
        logger.sysinfo('Exported to %s successful!' % (filename))

def print_dic(data):
    if 'url' in data.keys() and data['url'] != None and data['url'] != '':
        message = data['url']
    else:
        address = data['target_host'] + ":" + str(data['target_port']) if data['target_port'] != 0 else data['target_host']
        message = address

    if len(data['res']) == 0:
        msg = '[{0}] [{1}]'.format(data['module_name'], message)
        if data['flag'] == 1:
            logger.success(msg)
        elif data['flag'] == -1:
            logger.error(msg)
        else:
            logger.warning(msg)

    for res in data['res']:
        info = res['info'] if 'info' in res.keys() else ""
        key = res['key'] if 'key' in res.keys() else ""
        msg = '[{0}] [{1}]: {2}\t[{3}]'.format(data['module_name'], message, info, key)

        if data['flag'] == 1:
            logger.success(msg)
        elif data['flag'] == -1:
            logger.error(msg)
        else:
            logger.warning(msg)


def single_time_warn_message(message):  # Cross-linked function
    sys.stdout.write(message)
    sys.stdout.write("\n")
    sys.stdout.flush()


def data_to_stdout(data, bold=False):
    sys.stdout.write(set_color(data, bold))
    try:
        sys.stdout.flush()
    except IOError:
        pass


def set_color(message, bold=False):
    retVal = message
    if message and getattr(logger.console_handler, "is_tty", False):  # colorizing handler
        if bold:
            retVal = colored(message, color=None, on_color=None, attrs=("bold",))
    return retVal