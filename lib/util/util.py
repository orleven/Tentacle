#!/usr/bin/env python
# -*- encoding: utf-8 -*-
# @author: orleven

import time
import random
import string
from openpyxl import Workbook
from datetime import datetime
from urllib.parse import urlparse
from lib.util.cipherutil import md5
from lib.util.cipherutil import base64pickle
from lib.util.cipherutil import base64unpickle

def get_time(timestamp: float = None):
    """获取当前时间"""

    if timestamp:
        return datetime.fromtimestamp(timestamp)
    else:
        return datetime.fromtimestamp(get_timestamp())



def get_timestamp():
    """获取时间戳"""

    return time.time()


def random_string(length=32):
    """生成随机字符串"""

    return ''.join([random.choice(string.ascii_letters + string.digits + '_@^$') for _ in range(length)])


def random_lowercase_digits(length=16):
    """生成随机字符串"""

    return ''.join([random.choice(string.ascii_lowercase + string.digits) for _ in range(length)])


def random_digits(length=4):
    """生成随机数字的字符串"""

    return ''.join([random.choice(string.digits) for _ in range(length)])


def random_int(length=4):
    """生成随机数字的int"""

    return random.randint(10 ** (length - 1), 10 ** length - 1)

def random_md5(length=32, ret_plain=False):
    """
    生成随机MD5键值对

    :param ret_plain: 返回明文
    :param length:指定明文长度
    :param hex:指定密文长度为32位
    :returns 原文，密文(32位或16位)
    """
    plain = random_string(length)
    cipher = md5(plain)

    if ret_plain:
        return [plain, cipher]
    else:
        return cipher

def random_ua():
    """返回随机user_agent"""

    user_agent = [
        'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; WOW64; Trident/5.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; Zune 4.0; InfoPath.3; MS-RTC LM 8; .NET4.0C; .NET4.0E)',
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.87 Safari/537.36",
        'Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US; rv:1.9.2.3) Gecko/20100401 Firefox/3.0.16 (.NET CLR 3.5.30729)',
        'Mozilla/5.0 (Macintosh; U; PPC Mac OS X; en-US) AppleWebKit/125.4 (KHTML, like Gecko, Safari) OmniWeb/v563.57',
        'Mozilla/5.0 (X11; U; Linux i686; en-US) AppleWebKit/534.7 (KHTML, like Gecko) Chrome/7.0.517.24 Safari/534.7',
        'Mozilla/5.0 (X11; U; Linux x86_64; en-US) AppleWebKit/534.16 SUSE/10.0.626.0 (KHTML, like Gecko) Chrome/10.0.626.0 Safari/534.16',
        'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_6; en-US) AppleWebKit/534.20 (KHTML, like Gecko) Chrome/11.0.672.2 Safari/534.20',
        'Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US) AppleWebKit/534.7 (KHTML, like Gecko) Chrome/7.0.514.0 Safari/534.7',
        'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_3; de-de) AppleWebKit/531.22.7 (KHTML, like Gecko) Version/4.0.5 Safari/531.22.7',
        'Mozilla/5.0 (Windows; U; Windows NT 5.2; en-US) AppleWebKit/534.4 (KHTML, like Gecko) Chrome/6.0.481.0 Safari/534.4',
        'Mozilla/5.0 (Windows; U; Windows NT 6.0; en-GB; rv:1.9.1b4) Gecko/20090423 Firefox/3.5b4 (.NET CLR 3.5.30729)',
        'Mozilla/5.0 (Windows; U; Windows NT 6.0; nb-NO) AppleWebKit/533.18.1 (KHTML, like Gecko) Version/5.0.2 Safari/533.18.5',
        'Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US; rv:1.9.1.16) Gecko/20101130 AskTbPLTV5/3.8.0.12304 Firefox/3.5.16 (.NET CLR 3.5.30729)',
        'Mozilla/5.0 (X11; U; Linux i686; de; rv:1.9.0.14) Gecko/2009082505 Red Hat/3.0.14-1.el5_4 Firefox/3.0.14',
        'Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US; rv:1.8.1b2) Gecko/20060821 Firefox/2.0b2',
        'Mozilla/5.0 (compatible; Baiduspider/2.0; +http://www.baidu.com/search/spider.html)',
        'Opera/9.80 (Macintosh; Intel Mac OS X; U; nl) Presto/2.6.30 Version/10.61',
        'Googlebot/2.1 (+http://www.google.com/bot.html)'
        'Mozilla/5.0 (iPhone; CPU iPhone OS 9_3_4 like Mac OS X) AppleWebKit/601.1.46 (KHTML, like Gecko) Mobile/13G35 QQ/6.5.3.410 V1_IPH_SQ_6.5.3_1_APP_A Pixel/750 Core/UIWebView NetType/2G Mem/117"',
    ]
    return random.choice(user_agent)


def ip_header():
    """返回特制的header"""

    headers = {
        "X-Forwarded-For": "127.0.0.1",
        "X-Forwarded": "127.0.0.1",
        "Forwarded-For": "127.0.0.1",
        "Forwarded": "127.0.0.1",
        "X-Requested-With": "127.0.0.1",
        "X-Forwarded-Proto": "127.0.0.1",
        "X-Forwarded-Host": "127.0.0.1",
        "X-remote-IP": "127.0.0.1",
        "X-remote-addr": "127.0.0.1",
        "True-Client-IP": "127.0.0.1",
        "X-Client-IP": "127.0.0.1",
        "Client-IP": "127.0.0.1",
        "X-Real-IP": "127.0.0.1",
        "Ali-CDN-Real-IP": "127.0.0.1",
        "Cdn-Src-Ip": "127.0.0.1",
        "Cdn-Real-Ip": "127.0.0.1",
        "CF-Connecting-IP": "127.0.0.1",
        "X-Cluster-Client-IP": "127.0.0.1",
        "WL-Proxy-Client-IP": "127.0.0.1",
        "Proxy-Client-IP": "127.0.0.1",
        "Fastly-Client-Ip": "127.0.0.1",
        "True-Client-Ip": "127.0.0.1",
        "X-Originating-IP": "127.0.0.1",
        "X-Host": "127.0.0.1",
        "X-Custom-IP-Authorization": "127.0.0.1",
        "CF-Connecting_IP": "127.0.0.1",
    }
    return headers

def parse_url(url):
    """解析url"""

    try:
        arr = urlparse(url)
        scheme = arr.scheme
        hostname = arr.hostname
        port = arr.port
        if port is None:
            if scheme == 'https':
                port = 443
            else:
                port = 80
        return scheme, hostname, port
    except:
        return None, None, None


def get_base_url(url):
    """解析baseurl"""

    scheme, hostname, port = parse_url(url)

    if scheme == 'http' and port == 80:
        return f"{scheme}://{hostname}/"

    elif scheme == 'https' and port == 443:
        return f"{scheme}://{hostname}/"

    return f"{scheme}://{hostname}:{port}/"


def get_time_str(date: datetime = None, fmt="%Y-%m-%d %H:%M:%S") -> str:
    """获取时间字符串"""

    if date:
        return datetime.strftime(date, fmt)
    else:
        return datetime.strftime(get_time(), fmt)

def serialize_object(object_):
    return base64pickle(object_)

def unserialize_object(value):
    return base64unpickle(value) if value else None


def byte2hex(data):
    return data.hex()

def hex2byte(data):
    return bytes.fromhex(data)


def bin2dec(string_num):
    return str(int(string_num, 2))

def bin2hex(string_num):
    return dec2hex(bin2dec(string_num))

def dec2bin(string_num):
    base = [str(x) for x in range(10)] + [chr(x) for x in range(ord('A'), ord('A') + 6)]
    num = int(string_num)
    mid = []
    while True:
        if num == 0: break
        num,rem = divmod(num, 2)
        mid.append(base[rem])
    return ''.join([str(x) for x in mid[::-1]])

def dec2hex(string_num):
    base = [str(x) for x in range(10)] + [chr(x) for x in range(ord('A'), ord('A') + 6)]
    num = int(string_num)
    mid = []
    while True:
        if num == 0: break
        num,rem = divmod(num, 16)
        mid.append(base[rem])
    return ''.join([str(x) for x in mid[::-1]])


def hex2bin(string_num):
    return dec2bin(hex2dec(string_num.upper()))

def hex2dec(string_num):
    return str(int(string_num.upper(), 16))


def output_excal(datalines, filename):
    book = Workbook()
    ws = book.active
    i = 1
    titleList = []
    for line in datalines:
        i = i + 1
        for key in line:
            if key not in titleList:
                titleList.append(key)
                ws.cell(row=1, column=len(titleList)).value = key
            try:
                if line[key] == None or line[key] == '':
                    ws.cell(row=i, column=titleList.index(key) + 1).value = ""
                elif isinstance(line[key], int) or isinstance(line[key], str):
                    ws.cell(row=i, column=titleList.index(key) + 1).value = line[key]
                elif isinstance(line[key], bytes) :
                    ws.cell(row=i, column=titleList.index(key) + 1).value = str(line[key],'utf-8')
                elif isinstance(line[key], list) or isinstance(line[key], dict):
                    ws.cell(row=i, column=titleList.index(key) + 1).value = str(line[key])
                else:
                    ws.cell(row=i, column=titleList.index(key) + 1).value = "Types of printing are not supported."
            except:
                ws.cell(row=i, column=titleList.index(key) + 1).value = "Some error"
    book.save(filename)
